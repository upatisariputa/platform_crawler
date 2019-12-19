import os, sys, requests
from twitch_client_id import client_id
from http_data import url, headers
from openpyxl import Workbook, load_workbook
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from config.bj_num_datas_xl import get_id_from_xl

twitch_num_name_dic = get_id_from_xl('twitch.xlsx', 'twitch')

wb = load_workbook('twitch.xlsx')
ws = wb['twitch']

def get_bj_id(twitch_num_name_dic):
    twitch_bj_dic = {}
    row_num = 1
    for key, name in twitch_num_name_dic.items():
        if ws.cell(row_num, 3).value is None:
            res = requests.get(url + 'users?login=' + name[0], headers=headers).json()
            try:
                bj_id = res['data'][0]['id']
                ws.cell(row_num, 3, bj_id)
                wb.save('twitch.xlsx')
            except Exception:
                print('no : ', name)
                pass
        row_num = row_num + 1    

    return twitch_bj_dic


